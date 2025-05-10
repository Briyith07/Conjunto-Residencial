from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from models.usuario import Usuarios
from models.facturacion import Factura
from models.reserva import Reserva
from models.pqrs import PQRS
from models.casas import Apartamento, Torre, Casas
from models.roles import Roles
from models.datos_conjunto import DatosConjunto
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import os
from app import db
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from controllers.log import registrar_log
from controllers.configuracion import require_permission

reportes_bp = Blueprint('reportes', __name__)

def generar_pdf(titulo, datos, columnas, filename, usuario):
    carpeta_reportes = os.path.abspath(os.path.join("static", "reportes"))
    os.makedirs(carpeta_reportes, exist_ok=True)

    filepath = os.path.join(carpeta_reportes, filename)
    c = canvas.Canvas(filepath, pagesize=landscape(letter))
    width, height = landscape(letter)

    datos_conjunto = DatosConjunto.query.first()
    nombre_conjunto = datos_conjunto.nombre if datos_conjunto else "Conjunto Residencial"

    logo_path = os.path.abspath(os.path.join("static", "img", "logo.png"))
    if os.path.exists(logo_path):
        c.drawImage(logo_path, 50, height - 100, width=120, height=80, preserveAspectRatio=True, mask='auto')

    c.setFont("Helvetica-Bold", 16)
    c.drawString(220, height - 50, nombre_conjunto)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(300, height - 80, titulo)

    fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 115, f"Fecha de generación: {fecha_generacion}")
    c.drawString(50, height - 130, f"Generado por: {usuario}")

    y_position = height - 150
    data = [columnas] + datos

    table = Table(data, colWidths=[125] * len(columnas))
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 1, colors.black)
    ]))

    table.wrapOn(c, width, height)
    table.drawOn(c, 50, y_position - (len(datos) * 20))

    c.showPage()
    c.save()
    return filepath

def generar_excel(titulo, datos, columnas, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Agregar el título
    ws.merge_cells('A1:E1')
    ws['A1'] = titulo
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)

    # Escribir las cabeceras
    for col_num, column_title in enumerate(columnas, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}2"] = column_title
        ws[f"{col_letter}2"].font = openpyxl.styles.Font(bold=True)

    # Escribir los datos
    for row_num, row_data in enumerate(datos, 3):
        for col_num, data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = data

    # Guardar el archivo
    carpeta_reportes = os.path.abspath(os.path.join("static", "reportes"))
    os.makedirs(carpeta_reportes, exist_ok=True)
    filepath = os.path.join(carpeta_reportes, filename)
    wb.save(filepath)
    return filepath

@reportes_bp.route('/reportes', methods=['GET', 'POST'])
@login_required
@require_permission('Generar Informes')
def generar_reportes():
    torres = Torre.query.all()
    apartamentos = Apartamento.query.all()
    roles = Roles.query.filter(Roles.id != 99).all()  
    if request.method == 'POST':

        tipo_reporte = request.form.get('tipo_reporte')
        print(tipo_reporte)
        #Reservas
        filtro_reservas = request.form.get('filtro_reservas')
        print(filtro_reservas)
        #PQRS
        filtro_pqrs = request.form.get('filtro_pqrs')
        #Facturacion
        filtro_facturacion = request.form.get('filtro_facturacion')
        #Usuarios
        filtro_usuario_tipo = request.form.get('filtro_usuario_tipo')
        id_rol = request.form.get('rol_id')

        #Unidad Residencial
        id_torre  = request.form.get('torre_id')
        id_apartamento  = request.form.get('apartamento_id')

        #Fechas
        fecha_inicio  = request.form.get('fecha_inicio')
        fecha_fin  = request.form.get('fecha_fin')

        export = request.form.get('tipo_export')

# ------------------------------------------
        # RESERVAS APROBADAS
        if tipo_reporte == "reservas":
            if filtro_reservas =="confirmadas":
                titulo = "Reservas confirmadas"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Horario", "Espacio"]
                query = Reserva.query.filter_by(id_estado=2)

                if id_torre:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Reserva.fecha > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Reserva.fecha < fecha_fin)

                datos = [[r.usuario.nombre,r.usuario.casa.torre.nombre +" - "+ r.usuario.casa.apartamento.numero,  r.fecha, r.horario, r.espacios.nombre] for r in query.all()]
                filename_pdf = f"reservas_aprobadas_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"reservas_aprobadas_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)


            # ------------------------------------------
            elif filtro_reservas == "pendientes":
                titulo = "Reservas en espera"
                columnas = ["Usuario","Unidad Residencial", "Fecha", "Horario", "Espacio", "Estado"]
                query = Reserva.query.filter_by(id_estado=1)

                if id_torre:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Reserva.fecha > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Reserva.fecha < fecha_fin)

                datos = [[r.usuario.nombre,r.usuario.casa.torre.nombre +" - "+ r.usuario.casa.apartamento.numero,  r.fecha, r.horario, r.espacios.nombre] for r in query.all()]
                filename = f"reservas_en_espera_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"reservas_en_espera_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)


            # ------------------------------------------
            elif filtro_reservas == "canceladas":
                titulo = "Reservas Canceladas"
                columnas = ["Usuario","Unidad Residencial", "Fecha", "Horario", "Espacio", "Estado"]
                query = Reserva.query.filter_by(id_estado=3)

                if id_torre:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Reserva.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Reserva.fecha > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Reserva.fecha < fecha_fin)

                datos = [[r.usuario.nombre,r.usuario.casa.torre.nombre  +" - "+ r.usuario.casa.apartamento.numero,  r.fecha, r.horario, r.espacios.nombre] for r in query.all()]
                filename = f"reservas_canceladas_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"reservas_canceladas_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)


        # ------------------------------------------
        elif tipo_reporte == "pqrs":
            if filtro_pqrs == "registradas":
                titulo = "PQRS registradas"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
                query = PQRS.query.filter_by(id_estado=1)

                if id_torre:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(PQRS.fecha_creacion > fecha_inicio)
                if fecha_fin:
                    query = query.filter(PQRS.fecha_creacion < fecha_fin)

                datos = [[p.usuario.nombre,p.usuario.casa.torre.nombre  +" - "+  p.usuario.casa.apartamento.numero, p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre] for p in query.all()]
                filename = f"pqrs_registradas_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"pqrs_registradas_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)


            # ------------------------------------------
            elif filtro_pqrs == "en_proceso":
                titulo = "PQRS en proceso"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
                query = PQRS.query.filter_by(id_estado=2)

                if id_torre:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(PQRS.fecha_creacion > fecha_inicio)
                if fecha_fin:
                    query = query.filter(PQRS.fecha_creacion < fecha_fin)

                datos = [[p.usuario.nombre,p.usuario.casa.torre.nombre  +" - "+  p.usuario.casa.apartamento.numero, p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre] for p in query.all()]
                filename = f"pqrs_en_proceso_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"pqrs_en_proceso_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)


            # ------------------------------------------
            elif filtro_pqrs == "finalizadas":
                titulo = "PQRS Finalizadas"
                columnas = ["Usuario", "Unidad Residencial", "Fecha", "Estado", "Tipo"]
                query = PQRS.query.filter_by(id_estado=3)

                if id_torre:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(PQRS.usuario).join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(PQRS.fecha_creacion > fecha_inicio)
                if fecha_fin:
                    query = query.filter(PQRS.fecha_creacion < fecha_fin)

                datos = [[p.usuario.nombre,p.usuario.casa.torre.nombre  +" - "+  p.usuario.casa.apartamento.numero, p.fecha_creacion, p.estado.nombre, p.tipo_p.nombre] for p in query.all()]
                filename = f"pqrs_finalizadas_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"pqrs_finalizadas_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)

        # ------------------------------------------
        elif tipo_reporte == "facturacion":
            if filtro_facturacion == "en_mora":
                titulo = "Residentes en mora"
                columnas = ["Usuario", "Unidad Residencial", "Mes", "Año", "Total"]
                query = Factura.query.filter_by(estado="Pendiente")
                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Factura.fecha_emision > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Factura.fecha_emision < fecha_fin)

                datos = [[f.usuario_rel.nombre,f.usuario.casa.torre.nombre  +" - "+  f.usuario.casa.apartamento.numero, f.mes, f.year, f.total] for f in query.all()]
                filename = f"residentes_en_mora_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"residentes_en_mora_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)

            # ------------------------------------------
            elif filtro_facturacion == "al_dia":
                titulo = "Residentes al día"
                columnas = ["Usuario", "Unidad Residencial", "Mes", "Año", "Total"]
                query = Factura.query.filter_by(estado="Aprobado")

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Factura.fecha_emision > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Factura.fecha_emision < fecha_fin)


                datos = [[f.usuario_rel.nombre,f.usuario.casa.torre.nombre  +" - "+  f.usuario.casa.apartamento.numero, f.mes, f.year, f.total] for f in query.all()]
                filename = f"residentes_al_dia_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"residentes_al_dia_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)

            elif filtro_facturacion == "por_unidad":
                titulo = "Reporte por Residente"
                columnas = ["Usuario", "Unidad Residencial", "Mes", "Año", "Total"]

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)
                if fecha_inicio:
                    query = query.filter(Factura.fecha_emision > fecha_inicio)
                if fecha_fin:
                    query = query.filter(Factura.fecha_emision < fecha_fin)


                datos = [[f.usuario_rel.nombre,f.usuario.casa.torre.nombre  +" - "+  f.usuario.casa.apartamento.numero, f.mes, f.year, f.total] for f in Factura.query.all()]
                filename = f"residente_informe_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"residentes_informe_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)
            
        elif tipo_reporte == "usuarios":
            if filtro_usuario_tipo == "habilitados":
                titulo = "Reporte Usuarios Habilitados"
                columnas = ["Usuario", "Unidad Residencial", "Identificacion","Correo", "Telefono", "Estado"]
                query = Usuarios.query.filter_by(estado=1)

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)


                datos = [[u.nombre,u.casa.torre.nombre  +" - "+  u.casa.apartamento.numero, u.identificacion, u.email, u.telefono, "Habilitado"] for u in query.all()]
                filename = f"usuarios_habilitados_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"usuarios_habilitados_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)
            
            elif filtro_usuario_tipo == "inhabilitados":
                titulo = "Reporte Usuarios Inhabilitados"
                columnas = ["Usuario", "Unidad Residencial", "Identificacion","Correo", "Telefono", "Estado"]
                query = Usuarios.query.filter_by(estado=2)

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)

                datos = [[u.nombre,u.casa.torre.nombre  +" - "+  u.casa.apartamento.numero, u.identificacion, u.email, u.telefono, "Inhabilitado"] for u in query.all()]
                filename = f"usuarios_inhabilitados_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"usuarios_inhabilitados_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                filepath_excel = generar_excel(titulo, datos, columnas, filename_excel)

                return send_file(filepath_pdf, as_attachment=True)
            
            elif filtro_usuario_tipo == "por_rol":
                titulo = "Reporte Usuarios Por Rol"
                columnas = ["Usuario", "Unidad Residencial", "Identificacion","Correo", "Telefono", "Rol"]
                query = Usuarios.query.filter_by(id_rol==id_rol)

                if id_torre:
                    query = query.join(Usuarios.casa).filter(Casas.id_torre == id_torre)
                if id_apartamento:
                    query = query.join(Usuarios.casa).filter(Casas.id_apartamento == id_apartamento)


                datos = [[u.nombre,u.casa.torre.nombre  +" - "+  u.casa.apartamento.numero, u.identificacion, u.email, u.telefono, u.rol.nombre] for u in query.all()]
                filename = f"usuarios_por_rol_{datetime.now().strftime('%Y%m%d')}.pdf"
                filename_excel = f"usuarios_por_rol_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                  # Generar los archivos
                if export == "pdf":
                    filepath_pdf = generar_pdf(titulo, datos, columnas, filename_pdf, Usuarios)
                    return send_file(filepath_pdf, as_attachment=True)
                elif export == "excel":
                    filepath = generar_excel(titulo, datos, columnas, filename_excel)
                    return send_file(filepath, as_attachment=True)
                else:
                    flash("Seleccione un tipo de exportacion válido.", "danger")
                    return redirect(url_for('reportes.generar_reportes'))

                

        else:
            flash("Seleccione un tipo de reporte válido.", "danger")
            return redirect(url_for('reportes.generar_reportes'))
 
        registrar_log(current_user.id,"Reportes", "Se Genero Reporte "+str(titulo))


        if export == "pdf":
            filepath = generar_pdf(titulo, datos, columnas, filename, current_user.nombre)
        elif export == "excel":
            filepath = generar_excel(titulo, datos, columnas, filename_excel)
        else:
            flash("Seleccione un tipo de exportacion válido.", "danger")
            return redirect(url_for('reportes.generar_reportes'))
            
        return send_file(filepath, as_attachment=True)

    return render_template('reportes/reportes.html', torres=torres, apartamentos=apartamentos, roles=roles)